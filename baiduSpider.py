# -*- coding: utf-8 -*-  
__author__ = 'Mr.Q'

from selenium import webdriver
import time, re
from openpyxl import load_workbook,Workbook
import configparser
import os

def get_conf(conf_name):
    '''
    获取配置文件内容
    :param conf_name:配置名称
    :return:配置内容
    '''
    curpath = os.path.dirname(os.path.realpath(__file__))
    os.path.join(curpath, 'conf.ini')
    conf = configparser.ConfigParser()
    conf.read('conf.ini')
    conf_info = conf.get('myconf', conf_name)
    return conf_info

def get_web_info(driver):
    '''
    爬去页面信息，并实现下一页跳转。当无跳转按钮时结束。
    :param driver: 浏览器实体
    :return:爬取信息列表
    '''
    web_info = []
    # 获取初始页面，用于之后跳转
    base_page = driver.current_window_handle
    paper_list = driver.find_elements_by_class_name('res_t')
    for i in paper_list:
        paper_title = i.text
        i.find_element_by_tag_name('a').click()
        # 获得当前打开的第一个窗口句柄
        window_1 = driver.current_window_handle
        # 获得当前打开的所有窗口的句柄
        windows = driver.window_handles
        # 切换到当前最新的窗口
        for current_window in windows:
            if current_window != window_1:
                driver.switch_to.window(current_window)
        try:
            journal_title = driver.find_element_by_class_name('journal_title').text
            key_word = driver.find_element_by_class_name('kw_main').text
            author = driver.find_element_by_class_name('author_text').text
            # print(paper_title, journal_title, key_word, author)
            web_info.append([paper_title, journal_title, key_word, author])
            time.sleep(1)
            driver.close()
        except:
            print('详情页错误', driver.current_url)
            driver.close()
        driver.switch_to.window(base_page)
    # 查看是否有下一页
    try:
        next_page = driver.find_element_by_class_name('c-icon-page-next')
        next_page.click()
        print('next page')
        time.sleep(1)
    except:
        print('finish!!!')
        driver.close()
        return web_info
    web_info += get_web_info(driver)
    # driver.close()
    return web_info

def get_name_list(file_loc):
    '''
    从excel中获取作者和作者对应的网址
    :param file_loc:excel文件地址，在配置文件中配置
    :return:作者和作者对应的网址
    '''
    workbook = load_workbook(file_loc)
    sheet1 = workbook.get_sheet_by_name('Sheet1')
    names = [i.value for i in sheet1['A']]
    sites = [i.value for i in sheet1['B']]
    name_site = [names, sites]
    # print(name_site)
    return name_site



file_loc = get_conf('file_loc')
name_list = get_name_list(file_loc)
# url = get_conf('url')
# print(url)
wb = Workbook()
sheet = wb.active

for j in range(len(name_list[0])):
    init_driver = webdriver.Chrome()
    name = name_list[0][j]
    url = name_list[1][j]
    print(name, '开始')
    init_driver.get(url)
    # 访问出现空白页面时重新加载
    while init_driver.find_elements_by_class_name('res_t') == []:
        print(name, '页面空白刷新')
        init_driver.get(url)
    info = get_web_info(init_driver)
    # print(info)
    sheet.title = "Sheet1"
    for i in info:
        i += [name]
        sheet.append(i)
wb.save(r'F:/baiduSpider/paper_list.xlsx')
